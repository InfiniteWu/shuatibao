"""
Excel → 刷题宝JSON 自动转换脚本

约定格式（4列，无表头）：
  第1列：题型     （单选题 / 多选题 / 判断题）
  第2列：题目内容
  第3列：可选项   （用 ; 分隔，如 "选项A;选项B;选项C;选项D"）
  第4列：答案     （字母，如 A 或 ABC）

用法：
  python xlsx2json.py 题库文件.xlsx
  python xlsx2json.py 题库文件.xlsx 输出名

输出：
  题库文件.json（或指定的输出名.json）
"""

import json
import sys
import os
import openpyxl

TYPE_MAP = {
    '单选题': 'single',
    '多选题': 'multiple',
    '判断题': 'truefalse',
}


def answer_to_index(s, n_opts):
    """字母答案转索引：A→0, B→1, ABC→[0,1,2]"""
    if not isinstance(s, str):
        return None
    indices = [ord(ch) - 65 for ch in s.strip().upper() if 'A' <= ch <= 'Z']
    if not indices:
        return None
    indices = [i for i in indices if i < n_opts]
    if not indices:
        return None
    return indices[0] if len(indices) == 1 else sorted(indices)


def convert(filepath):
    wb = openpyxl.load_workbook(filepath)

    all_questions = []
    stats = {'single': 0, 'multiple': 0, 'truefalse': 0}
    skipped = 0

    for ws in wb.worksheets:
        for r in range(1, ws.max_row + 1):
            # 读取4列
            qtype_raw = ws.cell(r, 1).value
            stem = ws.cell(r, 2).value
            opts_raw = ws.cell(r, 3).value
            ans_raw = ws.cell(r, 4).value

            # 跳过空行
            if not qtype_raw or not stem:
                continue

            # 跳过表头行
            qtype_str = str(qtype_raw).strip()
            if qtype_str not in TYPE_MAP:
                skipped += 1
                continue

            qtype = TYPE_MAP[qtype_str]

            # 解析选项
            options = [o.strip() for o in str(opts_raw or '').split(';') if o.strip()]
            if qtype == 'truefalse':
                options = options[:2] if len(options) >= 2 else ['对', '错']
            if len(options) < 2:
                skipped += 1
                continue

            # 解析答案
            answer = answer_to_index(str(ans_raw or '').strip(), len(options))
            if answer is None:
                skipped += 1
                continue
            if qtype == 'multiple' and not isinstance(answer, list):
                answer = [answer]

            all_questions.append({
                'type': qtype,
                'stem': str(stem).strip(),
                'options': options,
                'answer': answer,
                'explanation': '',
            })
            stats[qtype] += 1

    return all_questions, stats, skipped


def main():
    if len(sys.argv) < 2:
        print('用法: python xlsx2json.py <Excel文件路径> [输出名称]')
        print('示例: python xlsx2json.py 题库.xlsx')
        print('      python xlsx2json.py 题库.xlsx 我的题库')
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.isfile(filepath):
        print(f'错误：文件不存在 — {filepath}')
        sys.exit(1)

    # 输出名
    if len(sys.argv) >= 3:
        bank_name = sys.argv[2]
        out_name = sys.argv[2]
    else:
        bank_name = os.path.splitext(os.path.basename(filepath))[0]
        out_name = bank_name

    print(f'读取: {filepath}')
    questions, stats, skipped = convert(filepath)

    if not questions:
        print('错误：没有找到有效题目，请检查格式是否正确。')
        print('      第1列=题型，第2列=题目，第3列=选项(;分隔)，第4列=答案(A/B/C)')
        sys.exit(1)

    result = {
        'bankName': bank_name,
        'description': '单选{} 多选{} 判断{} 共{}题'.format(
            stats['single'], stats['multiple'], stats['truefalse'], len(questions)
        ),
        'questions': questions,
    }

    out_path = f'{out_name}.json'
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f'输出: {out_path}')
    print(f'  单选: {stats["single"]}')
    print(f'  多选: {stats["multiple"]}')
    print(f'  判断: {stats["truefalse"]}')
    print(f'  总计: {len(questions)} 题')
    if skipped:
        print(f'  跳过: {skipped} 行（表头或格式不符）')


if __name__ == '__main__':
    main()
