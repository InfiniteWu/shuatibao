"""安环2026.xlsx -> 刷题宝JSON"""
import json
import openpyxl

wb = openpyxl.load_workbook('安环2026.xlsx')
ws = wb.active

TYPE_MAP = {'单选题': 'single', '多选题': 'multiple', '判断题': 'truefalse'}


def answer_to_index(s, n_opts):
    """A/B/C -> 0/1/2, ABC -> [0,1,2]"""
    if not isinstance(s, str):
        return None
    indices = [ord(ch) - 65 for ch in s.strip() if 'A' <= ch <= 'Z']
    if not indices:
        return None
    indices = [i for i in indices if i < n_opts]
    if not indices:
        return None
    return indices[0] if len(indices) == 1 else indices


questions = []
stats = {'single': 0, 'multiple': 0, 'truefalse': 0}

for r in range(2, ws.max_row + 1):
    t = ws.cell(r, 2).value
    if t not in TYPE_MAP:
        continue

    qtype = TYPE_MAP[t]
    stem = str(ws.cell(r, 3).value or '').strip()
    options = [o.strip() for o in str(ws.cell(r, 4).value or '').split(';') if o.strip()]
    ans_raw = ws.cell(r, 5).value

    if len(options) < 2 or not ans_raw:
        continue

    if qtype == 'truefalse':
        options = options[:2]

    answer = answer_to_index(str(ans_raw), len(options))
    if answer is None:
        continue

    if qtype == 'multiple' and not isinstance(answer, list):
        answer = [answer]

    questions.append({
        'type': qtype,
        'stem': stem,
        'options': options,
        'answer': answer,
        'explanation': '',
    })
    stats[qtype] += 1

result = {
    'bankName': '安环2026',
    'description': f'单选{stats["single"]} 多选{stats["multiple"]} 判断{stats["truefalse"]} 共{len(questions)}题',
    'questions': questions,
}

out = '安环2026.json'
with open(out, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

with open('_convert_report.txt', 'w', encoding='utf-8') as f:
    f.write(f'输出: {out}\n')
    f.write(f'单选: {stats["single"]}\n')
    f.write(f'多选: {stats["multiple"]}\n')
    f.write(f'判断: {stats["truefalse"]}\n')
    f.write(f'总计: {len(questions)}\n')

    f.write('\n--- 前5题 ---\n')
    for q in questions[:5]:
        f.write(f"  [{q['type']}] {q['stem'][:60]}... answer={q['answer']}\n")

    f.write('\n--- 多选题样本 ---\n')
    for q in questions:
        if q['type'] == 'multiple':
            f.write(f"  answer={q['answer']} options={q['options']}\n")
            break

    f.write('\n--- 判断题样本 ---\n')
    for q in questions:
        if q['type'] == 'truefalse':
            f.write(f"  answer={q['answer']} options={q['options']}\n")
            break

print('Done')
